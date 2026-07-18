#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

def verify_updated_file(file_path):
    """验证更新后的Excel文件"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook["配置详细说明"]

        print(f"=== 验证更新后的Excel文件 ===")
        print(f"工作表最大行: {sheet.max_row}")
        print(f"工作表最大列: {sheet.max_column}")
        print()

        # 读取新增的数据（从第44行开始）
        print("=== 新增的Redis配置 ===")
        for row_num in range(44, 57):
            row_data = []
            for col in sheet[row_num]:
                row_data.append(col.value)
            print(f"Row {row_num}: {row_data}")

        print()
        print("=== 新增的Kafka配置 ===")
        for row_num in range(58, 69):
            row_data = []
            for col in sheet[row_num]:
                row_data.append(col.value)
            print(f"Row {row_num}: {row_data}")

        print()
        print("=== 新增的Zookeeper配置 ===")
        for row_num in range(70, 80):
            row_data = []
            for col in sheet[row_num]:
                row_data.append(col.value)
            print(f"Row {row_num}: {row_data}")

        print()
        print("=== 新增的AP应用配置 ===")
        for row_num in range(81, 92):
            row_data = []
            for col in sheet[row_num]:
                row_data.append(col.value)
            print(f"Row {row_num}: {row_data}")

        return True

    except Exception as e:
        print(f"验证文件时出错: {e}")
        return False

if __name__ == "__main__":
    file_path = "资源申请模板-资源申请20260520V0.1_updated.xlsx"
    verify_updated_file(file_path)