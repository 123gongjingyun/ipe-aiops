#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

def add_service_configurations(file_path):
    """为Redis、Kafka、Zookeeper和AP添加配置详细说明"""

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook["配置详细说明"]

        # 获取最后一个有数据的行
        max_row = sheet.max_row

        print(f"当前工作表最大行: {max_row}")

        # 从第44行开始添加新数据
        start_row = 44

        # Redis单节点配置
        redis_data = [
            ('单节点redis', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('CPU/内存', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '最大连接数', '每秒操作数(QPS)', '内存使用率', '适用场景', '适用用户规模', None, None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G 普通云盘', '500-1000', '40-90', '1000-5000', '10000-50000', '<80%', '开发测试', '日活<100', None, None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G 高效云盘', 2600, 115, '5000-10000', '50000-100000', '<80%', '测试环境', '日活<500', None, None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G SSD云盘', '3000-4800', '125-180', '10000-20000', '100000-200000', '<80%', '小型生产', '日活<2000', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G SSD云盘', 4800, 140, '20000-50000', '200000-500000', '<80%', '中型生产', '日活<1万', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G 增强型SSD', '8000+', '180+', '50000-100000', '500000-1000000', '<80%', '中型生产', '日活<5万', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G ESSD PL1', 6800, 170, '100000-150000', '1000000-1500000', '<80%', '大型生产', '日活<20万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G SSD云盘', 4800, 140, '150000-300000', '1500000-3000000', '<80%', '大型生产', '日活<50万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G 增强型SSD', '8000-50000', '350+', '300000-500000', '3000000-5000000', '<80%', '超大型生产', '日活<200万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G ESSD PL2', 25000, 260, '500000-1000000', '5000000-10000000', '<80%', '超大型生产', '日活<500万', None, None, None, None, None, None, None, None, None),
            ('16C32G', '单节点', '40G ESSD PL1', '1T ESSD PL2', 50000, 500, '1000000-2000000', '10000000-20000000', '<80%', '超大规模', '日活>500万', None, None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None)
        ]

        # Kafka单节点配置
        kafka_data = [
            ('单节点kafka', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('CPU/内存', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '消息吞吐量(msg/s)', '分区数量', '副本因子', 'Broker节点数', '适用场景', '适用用户规模', None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '200G 普通云盘', '500-1000', '40-90', '5000-10000', '10-50', '1-2', 1, '开发测试', '日活<100', None, None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '200G 高效云盘', 2600, 115, '10000-30000', '50-100', '2-3', 1, '测试环境', '日活<500', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G SSD云盘', 4800, 140, '30000-80000', '100-500', '2-3', 1, '小型生产', '日活<2000', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '500G 增强型SSD', '8000+', '180+', '80000-150000', '500-1000', '2-3', 1, '中型生产', '日活<5万', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '500G ESSD PL1', 6800, 170, '150000-300000', '1000-2000', '3', 1, '大型生产', '日活<20万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '1T SSD云盘', 4800, 140, '300000-500000', '2000-5000', '3', 1, '大型生产', '日活<50万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '1T 增强型SSD', '8000-50000', '350+', '500000-1000000', '5000-10000', '3', 1, '超大型生产', '日活<200万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '1T ESSD PL2', 25000, 260, '1000000-2000000', '10000-20000', '3', 1, '超大型生产', '日活<500万', None, None, None, None, None, None, None, None, None),
            ('16C32G', '单节点', '40G ESSD PL1', '2T ESSD PL2', 50000, 500, '2000000-5000000', '20000-50000', '3', 1, '超大规模', '日活>500万', None, None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None)
        ]

        # Zookeeper单节点配置
        zookeeper_data = [
            ('单节点zookeeper', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('CPU/内存', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '客户端连接数', '写操作QPS', '读操作QPS', '集群节点数', '适用场景', '适用用户规模', None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G 普通云盘', '500-1000', '40-90', '100-500', '1000-5000', '5000-20000', 1, '开发测试', '日活<100', None, None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G 高效云盘', 2600, 115, '500-1000', '5000-10000', '20000-50000', 1, '测试环境', '日活<500', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G SSD云盘', 4800, 140, '1000-3000', '10000-30000', '50000-150000', 1, '小型生产', '日活<2000', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G 增强型SSD', '8000+', '180+', '3000-8000', '30000-80000', '150000-400000', 1, '中型生产', '日活<5万', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G ESSD PL1', 6800, 170, '8000-15000', '80000-150000', '400000-800000', 1, '大型生产', '日活<20万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G SSD云盘', 4800, 140, '15000-30000', '150000-300000', '800000-2000000', 1, '大型生产', '日活<50万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G 增强型SSD', '8000-50000', '350+', '30000-80000', '300000-800000', '2000000-5000000', 1, '超大型生产', '日活<200万', None, None, None, None, None, None, None, None, None),
            ('16C32G', '单节点', '40G ESSD PL1', '1T ESSD PL2', 50000, 500, '80000-150000', '800000-1500000', '5000000-10000000', 1, '超大型生产', '日活<500万', None, None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None)
        ]

        # AP应用配置
        ap_data = [
            ('单节点AP应用', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('CPU/内存', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '并发用户数', '每秒请求数(req/s)', '响应时间(ms)', '吞吐量(req/s)', '适用场景', '适用用户规模', None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G 普通云盘', '500-1000', '40-90', '50-200', '20-100', '<500', '100-500', '开发测试', '日活<100', None, None, None, None, None, None, None, None, None),
            ('2C4G', '单节点', '40G 普通云盘', '100G 高效云盘', 2600, 115, '200-500', '100-300', '<300', '500-1500', '测试环境', '日活<500', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G SSD云盘', 4800, 140, '500-1500', '300-800', '<200', '1500-4000', '小型生产', '日活<2000', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G 增强型SSD', '8000+', '180+', '1500-4000', '800-2000', '<150', '4000-10000', '中型生产', '日活<1万', None, None, None, None, None, None, None, None, None),
            ('4C8G', '单节点', '40G SSD云盘', '200G ESSD PL1', 6800, 170, '4000-8000', '2000-5000', '<100', '10000-25000', '大型生产', '日活<5万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G SSD云盘', 4800, 140, '8000-20000', '5000-12000', '<80', '25000-60000', '大型生产', '日活<20万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G 增强型SSD', '8000-50000', '350+', '20000-50000', '12000-30000', '<50', '60000-150000', '超大型生产', '日活<50万', None, None, None, None, None, None, None, None, None),
            ('8C16G', '单节点', '40G SSD云盘', '500G ESSD PL2', 25000, 260, '50000-100000', '30000-80000', '<30', '150000-300000', '超大型生产', '日活<200万', None, None, None, None, None, None, None, None, None),
            ('16C32G', '单节点', '40G ESSD PL1', '1T ESSD PL2', 50000, 500, '100000-200000', '80000-200000', '<20', '300000-600000', '超大规模', '日活<500万', None, None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None)
        ]

        # 合并所有数据
        all_new_data = redis_data + kafka_data + zookeeper_data + ap_data

        # 复制现有样式
        def get_cell_styles(cell):
            return {
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format
            }

        # 获取样式的参考行（使用第20行作为参考，因为它是RabbitMQ的标题行）
        reference_row = 20
        title_styles = {}
        header_styles = {}
        data_styles = {}

        for col in range(1, 21):  # A到T列
            title_styles[col] = get_cell_styles(sheet.cell(row=1, column=col))
            header_styles[col] = get_cell_styles(sheet.cell(row=20, column=col))
            data_styles[col] = get_cell_styles(sheet.cell(row=21, column=col))

        # 添加新数据
        current_row = start_row
        for row_data in all_new_data:
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)

                # 应用样式
                if current_row == start_row or '单节点' in str(row_data[0]) or '集群' in str(row_data[0]) or 'AP应用' in str(row_data[0]):
                    # 标题行样式
                    styles = title_styles.get(col, {})
                elif any(isinstance(v, str) and v in ['CPU/内存', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '适用场景', '适用用户规模',
                                                      '并发连接数', '消息吞吐量(msg/s)', '队列数量', '每秒操作数(QPS)', '内存使用率', '最大连接数',
                                                      '消息吞吐量(msg/s)', '分区数量', '副本因子', 'Broker节点数', '客户端连接数', '写操作QPS', '读操作QPS', '集群节点数',
                                                      '并发用户数', '每秒请求数(req/s)', '响应时间(ms)', '吞吐量(req/s)'] for v in row_data):
                    # 表头行样式
                    styles = header_styles.get(col, {})
                else:
                    # 数据行样式
                    styles = data_styles.get(col, {})

                # 应用样式到单元格
                if 'font' in styles:
                    cell.font = styles['font']
                if 'fill' in styles:
                    cell.fill = styles['fill']
                if 'border' in styles:
                    cell.border = styles['border']
                if 'alignment' in styles:
                    cell.alignment = styles['alignment']
                if 'number_format' in styles:
                    cell.number_format = styles['number_format']

            current_row += 1

        # 保存文件
        output_file = "资源申请模板-资源申请20260520V0.1_updated.xlsx"
        workbook.save(output_file)
        print(f"✅ 文件已更新并保存为: {output_file}")
        print(f"📊 新增了 {len(all_new_data)} 行数据")
        print(f"📍 数据从第 {start_row} 行开始")

        return True

    except Exception as e:
        print(f"❌ 处理文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "资源申请模板-资源申请20260520V0.1.xlsx"
    add_service_configurations(file_path)