#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

def add_kafka_zookeeper_cluster_configurations(file_path):
    """添加Kafka集群、Zookeeper集群和组合架构配置"""

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook["配置详细说明"]

        # Kafka 3节点集群架构配置
        kafka_cluster_data = [
            ('Kafka集群3节点', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('Broker节点配置', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '集群消息吞吐量(msg/s)', '集群分区数量', '副本因子', 'Broker节点数', '适用场景', '适用用户规模', None, None, None, None, None, None, None),
            ('2C4G x 3节点', '3节点集群', '40G SSD云盘 x3', '200G SSD云盘 x3', 4800, 140, '90000-240000', '300-1500', '2-3', 3, '小型集群', '日活<5000', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '3节点集群', '40G SSD云盘 x3', '500G SSD云盘 x3', 4800, 140, '240000-600000', '1500-6000', '2-3', 3, '中型集群', '日活<2万', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '3节点集群', '40G SSD云盘 x3', '500G 增强型SSD x3', '8000+', '180+', '600000-1500000', '6000-15000', '3', 3, '中型集群', '日活<10万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '3节点集群', '40G SSD云盘 x3', '1T SSD云盘 x3', 4800, 140, '1500000-3000000', '15000-30000', '3', 3, '大型集群', '日活<50万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '3节点集群', '40G SSD云盘 x3', '1T 增强型SSD x3', '8000-50000', '350+', '3000000-9000000', '30000-90000', '3', 3, '大型集群', '日活<200万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '3节点集群', '40G SSD云盘 x3', '1T ESSD PL2 x3', 25000, 260, '9000000-18000000', '90000-180000', '3', 3, '超大型集群', '日活<500万', None, None, None, None, None, None, None, None),
            ('16C32G x 3节点', '3节点集群', '40G ESSD PL1 x3', '2T ESSD PL2 x3', 50000, 500, '18000000-60000000', '180000-600000', '3', 3, '超大型集群', '日活>500万', None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        ]

        # Kafka + Zookeeper组合架构配置
        kafka_zookeeper_data = [
            ('Kafka+Zookeeper组合架构', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('Zookeeper节点配置', 'Kafka Broker节点配置', '架构类型', 'ZK系统盘', 'Kafka数据盘', '数据盘IOPS', '数据盘吞吐量(MB/s)', '集群消息吞吐量(msg/s)', '集群分区数量', '总节点数', '适用场景', '适用用户规模', None, None, None, None, None, None, None),
            ('2C4G x 3节点', '2C4G x 3节点', 'ZK3节点+Kafka3节点', '40G SSD云盘 x3', '200G SSD云盘 x3', 4800, 140, '60000-180000', '150-900', 6, '小型组合架构', '日活<3000', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '4C8G x 3节点', 'ZK3节点+Kafka3节点', '40G SSD云盘 x3', '500G SSD云盘 x3', 4800, 140, '180000-450000', '900-3000', 6, '中型组合架构', '日活<1万', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '4C8G x 3节点', 'ZK3节点+Kafka3节点', '40G SSD云盘 x3', '500G 增强型SSD x3', '8000+', '180+', '450000-1200000', '3000-9000', 6, '中型组合架构', '日活<5万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '8C16G x 3节点', 'ZK3节点+Kafka3节点', '40G SSD云盘 x3', '1T SSD云盘 x3', 4800, 140, '1200000-3000000', '9000-30000', 6, '大型组合架构', '日活<30万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '8C16G x 3节点', 'ZK3节点+Kafka3节点', '40G SSD云盘 x3', '1T 增强型SSD x3', '8000-50000', '350+', '3000000-7500000', '30000-75000', 6, '大型组合架构', '日活<150万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '8C16G x 3节点', 'ZK3节点+Kafka3节点', '40G SSD云盘 x3', '1T ESSD PL2 x3', 25000, 260, '7500000-15000000', '75000-150000', 6, '超大型组合架构', '日活<400万', None, None, None, None, None, None, None, None),
            ('16C32G x 3节点', '16C32G x 3节点', 'ZK3节点+Kafka3节点', '40G ESSD PL1 x3', '2T ESSD PL2 x3', 50000, 500, '15000000-45000000', '150000-450000', 6, '超大型组合架构', '日活>400万', None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        ]

        # Zookeeper 3节点集群架构配置
        zookeeper_cluster_data = [
            ('Zookeeper集群3节点', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('节点配置', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '集群客户端连接数', '集群写操作QPS', '集群读操作QPS', '集群节点数', '适用场景', '适用用户规模', None, None, None, None, None, None, None),
            ('2C4G x 3节点', '3节点集群', '40G SSD云盘 x3', '100G SSD云盘 x3', 4800, 140, '1500-3000', '15000-45000', '75000-225000', 3, '小型集群', '日活<2000', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '3节点集群', '40G SSD云盘 x3', '200G SSD云盘 x3', 4800, 140, '3000-9000', '45000-135000', '225000-675000', 3, '中型集群', '日活<1万', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '3节点集群', '40G SSD云盘 x3', '200G 增强型SSD x3', '8000+', '180+', '9000-24000', '135000-360000', '675000-1800000', 3, '中型集群', '日活<5万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '3节点集群', '40G SSD云盘 x3', '500G SSD云盘 x3', 4800, 140, '24000-60000', '360000-900000', '1800000-4500000', 3, '大型集群', '日活<20万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '3节点集群', '40G SSD云盘 x3', '500G 增强型SSD x3', '8000-50000', '350+', '60000-150000', '900000-2250000', '4500000-11250000', 3, '大型集群', '日活<100万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '3节点集群', '40G SSD云盘 x3', '500G ESSD PL2 x3', 25000, 260, '150000-300000', '2250000-4500000', '11250000-22500000', 3, '超大型集群', '日活<300万', None, None, None, None, None, None, None, None),
            ('16C32G x 3节点', '3节点集群', '40G ESSD PL1 x3', '1T ESSD PL2 x3', 50000, 500, '300000-600000', '4500000-9000000', '22500000-45000000', 3, '超大型集群', '日活>300万', None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        ]

        # 合并所有新数据
        all_new_data = kafka_cluster_data + kafka_zookeeper_data + zookeeper_cluster_data

        # 获取最后一个有数据的行
        max_row = sheet.max_row
        print(f"当前工作表最大行: {max_row}")

        # 从第111行开始添加新数据
        start_row = max_row + 1

        # 复制现有样式的函数
        def get_cell_styles(cell):
            return {
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format
            }

        # 获取样式的参考行
        title_styles = {}
        header_styles = {}
        data_styles = {}

        for col in range(1, 21):  # A到T列
            # 使用Kafka单节点的样式作为参考
            title_styles[col] = get_cell_styles(sheet.cell(row=57, column=col))
            header_styles[col] = get_cell_styles(sheet.cell(row=58, column=col))
            data_styles[col] = get_cell_styles(sheet.cell(row=59, column=col))

        # 添加新数据
        current_row = start_row
        for row_data in all_new_data:
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)

                # 应用样式
                if any(str(value).startswith(x) for x in ['Kafka集群', 'Kafka+Zookeeper', 'Zookeeper集群'] if value):
                    # 标题行样式
                    styles = title_styles.get(col, {})
                elif any(isinstance(v, str) and v in ['Broker节点配置', 'Zookeeper节点配置', 'Kafka Broker节点配置', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)',
                                                          '集群消息吞吐量(msg/s)', '集群分区数量', '副本因子', 'Broker节点数', '总节点数', '适用场景', '适用用户规模',
                                                          '集群客户端连接数', '集群写操作QPS', '集群读操作QPS', '集群节点数'] for v in row_data):
                    # 表头行样式
                    styles = header_styles.get(col, {})
                else:
                    # 数据行样式
                    styles = data_styles.get(col, {})

                # 应用样式到单元格
                if 'font' in styles:
                    cell.font = styles['font']
                if 'fill' in styles:
                    cell.fill = styles['fill']
                if 'border' in styles:
                    cell.border = styles['border']
                if 'alignment' in styles:
                    cell.alignment = styles['alignment']
                if 'number_format' in styles:
                    cell.number_format = styles['number_format']

            current_row += 1

        # 保存文件
        output_file = "资源申请模板-资源申请20260520V0.1_updated.xlsx"
        workbook.save(output_file)
        print(f"✅ Kafka和Zookeeper集群架构配置已添加")
        print(f"📊 新增了 {len(all_new_data)} 行数据")
        print(f"📍 数据从第 {start_row} 行开始")
        print(f"💾 文件已保存: {output_file}")

        return True

    except Exception as e:
        print(f"❌ 处理文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "资源申请模板-资源申请20260520V0.1_updated.xlsx"
    add_kafka_zookeeper_cluster_configurations(file_path)