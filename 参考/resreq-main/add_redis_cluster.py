#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

def add_redis_cluster_configurations(file_path):
    """为Redis添加集群架构配置：3主3从、一主2从3哨兵"""

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook["配置详细说明"]

        # 找到Redis单节点配置结束的位置（第56行是空行）
        # 在第56行后插入集群架构配置

        # Redis 3主3从集群架构配置
        redis_cluster_data = [
            ('Redis集群3主3从', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('主节点配置', '从节点配置', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '集群总连接数', '集群总QPS', '内存使用率', '分片数量', '适用场景', '适用用户规模', None, None, None, None, None, None, None),
            ('2C4G x 3节点', '2C4G x 3节点', '3主3从', '40G SSD云盘 x6', '100G SSD云盘 x6', 4800, 140, '30000-60000', '600000-1200000', '<80%', 3, '小型集群', '日活<1万', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '4C8G x 3节点', '3主3从', '40G SSD云盘 x6', '200G SSD云盘 x6', 4800, 140, '60000-150000', '1200000-3000000', '<80%', 3, '中型集群', '日活<5万', None, None, None, None, None, None, None, None),
            ('4C8G x 3节点', '4C8G x 3节点', '3主3从', '40G SSD云盘 x6', '200G 增强型SSD x6', '8000+', '180+', '150000-300000', '3000000-6000000', '<80%', 3, '中型集群', '日活<20万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '8C16G x 3节点', '3主3从', '40G SSD云盘 x6', '500G SSD云盘 x6', 4800, 140, '300000-600000', '6000000-12000000', '<80%', 3, '大型集群', '日活<50万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '8C16G x 3节点', '3主3从', '40G SSD云盘 x6', '500G 增强型SSD x6', '8000-50000', '350+', '600000-1500000', '12000000-30000000', '<80%', 3, '大型集群', '日活<200万', None, None, None, None, None, None, None, None),
            ('8C16G x 3节点', '8C16G x 3节点', '3主3从', '40G SSD云盘 x6', '500G ESSD PL2 x6', 25000, 260, '1500000-3000000', '30000000-60000000', '<80%', 3, '超大型集群', '日活<500万', None, None, None, None, None, None, None, None),
            ('16C32G x 3节点', '16C32G x 3节点', '3主3从', '40G ESSD PL1 x6', '1T ESSD PL2 x6', 50000, 500, '3000000-6000000', '60000000-120000000', '<80%', 3, '超大型集群', '日活>500万', None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        ]

        # Redis 一主2从3哨兵架构配置
        redis_sentinel_data = [
            ('Redis哨兵一主2从3哨兵', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
            ('主节点配置', '从节点配置', '哨兵节点配置', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)', '集群总连接数', '集群总QPS', '内存使用率', '适用场景', '适用用户规模', None, None, None, None, None, None, None),
            ('2C4G', '2C4G x 2节点', '1C2G x 3节点', '一主二从三哨兵', '40G SSD云盘 x3', '100G SSD云盘 x3', 4800, 140, '15000-30000', '300000-600000', '<80%', '小型高可用', '日活<5000', None, None, None, None, None, None, None, None),
            ('4C8G', '4C8G x 2节点', '1C2G x 3节点', '一主二从三哨兵', '40G SSD云盘 x3', '200G SSD云盘 x3', 4800, 140, '30000-75000', '600000-1500000', '<80%', '中型高可用', '日活<2万', None, None, None, None, None, None, None, None),
            ('4C8G', '4C8G x 2节点', '1C2G x 3节点', '一主二从三哨兵', '40G SSD云盘 x3', '200G 增强型SSD x3', '8000+', '180+', '75000-150000', '1500000-3000000', '<80%', '中型高可用', '日活<10万', None, None, None, None, None, None, None, None),
            ('8C16G', '8C16G x 2节点', '2C4G x 3节点', '一主二从三哨兵', '40G SSD云盘 x3', '500G SSD云盘 x3', 4800, 140, '150000-300000', '3000000-6000000', '<80%', '大型高可用', '日活<50万', None, None, None, None, None, None, None, None),
            ('8C16G', '8C16G x 2节点', '2C4G x 3节点', '一主二从三哨兵', '40G SSD云盘 x3', '500G 增强型SSD x3', '8000-50000', '350+', '300000-750000', '6000000-15000000', '<80%', '大型高可用', '日活<200万', None, None, None, None, None, None, None, None),
            ('8C16G', '8C16G x 2节点', '2C4G x 3节点', '一主二从三哨兵', '40G SSD云盘 x3', '500G ESSD PL2 x3', 25000, 260, '750000-1500000', '15000000-30000000', '<80%', '超大型高可用', '日活<500万', None, None, None, None, None, None, None, None),
            ('16C32G', '16C32G x 2节点', '2C4G x 3节点', '一主二从三哨兵', '40G ESSD PL1 x3', '1T ESSD PL2 x3', 50000, 500, '1500000-3000000', '30000000-60000000', '<80%', '超大型高可用', '日活>500万', None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        ]

        # 合并所有新数据
        all_new_data = redis_cluster_data + redis_sentinel_data

        # 获取最后一个有数据的行
        max_row = sheet.max_row
        print(f"当前工作表最大行: {max_row}")

        # 从第92行开始添加新数据（在现有数据之后）
        start_row = max_row + 1

        # 复制现有样式的函数
        def get_cell_styles(cell):
            return {
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format
            }

        # 获取样式的参考行
        title_styles = {}
        header_styles = {}
        data_styles = {}

        for col in range(1, 21):  # A到T列
            # 使用Redis单节点的样式作为参考
            title_styles[col] = get_cell_styles(sheet.cell(row=44, column=col))
            header_styles[col] = get_cell_styles(sheet.cell(row=45, column=col))
            data_styles[col] = get_cell_styles(sheet.cell(row=46, column=col))

        # 添加新数据
        current_row = start_row
        for row_data in all_new_data:
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)

                # 应用样式
                if any(str(value).startswith(x) for x in ['Redis集群', 'Redis哨兵', '单节点redis', '单节点kafka'] if value):
                    # 标题行样式
                    styles = title_styles.get(col, {})
                elif any(isinstance(v, str) and v in ['主节点配置', '从节点配置', '哨兵节点配置', '架构类型', '系统盘配置', '数据盘配置', '数据盘IOPS', '数据盘吞吐量(MB/s)',
                                                          '集群总连接数', '集群总QPS', '内存使用率', '分片数量', '适用场景', '适用用户规模',
                                                          'CPU/内存', '架构类型', '最大连接数', '每秒操作数(QPS)'] for v in row_data):
                    # 表头行样式
                    styles = header_styles.get(col, {})
                else:
                    # 数据行样式
                    styles = data_styles.get(col, {})

                # 应用样式到单元格
                if 'font' in styles:
                    cell.font = styles['font']
                if 'fill' in styles:
                    cell.fill = styles['fill']
                if 'border' in styles:
                    cell.border = styles['border']
                if 'alignment' in styles:
                    cell.alignment = styles['alignment']
                if 'number_format' in styles:
                    cell.number_format = styles['number_format']

            current_row += 1

        # 保存文件
        output_file = "资源申请模板-资源申请20260520V0.1_updated.xlsx"
        workbook.save(output_file)
        print(f"✅ Redis集群架构配置已添加")
        print(f"📊 新增了 {len(all_new_data)} 行数据")
        print(f"📍 数据从第 {start_row} 行开始")
        print(f"💾 文件已保存: {output_file}")

        return True

    except Exception as e:
        print(f"❌ 处理文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "资源申请模板-资源申请20260520V0.1_updated.xlsx"
    add_redis_cluster_configurations(file_path)