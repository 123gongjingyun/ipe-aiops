#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys

def read_excel_sheet(file_path, sheet_name):
    """读取Excel文件的指定sheet"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # 检查sheet是否存在
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in workbook.")
            print(f"Available sheets: {workbook.sheetnames}")
            return None

        sheet = workbook[sheet_name]

        print(f"=== Sheet: {sheet_name} ===")
        print(f"Dimensions: {sheet.dimensions}")
        print()

        # 读取所有数据
        data = []
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            data.append(row)
            # 打印每一行
            print(f"Row {row_num}: {row}")

        return data

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

if __name__ == "__main__":
    file_path = "资源申请模板-资源申请20260520V0.1.xlsx"
    sheet_name = "配置详细说明"

    data = read_excel_sheet(file_path, sheet_name)